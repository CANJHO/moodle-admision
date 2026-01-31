# actas_presentacion.py
# Genera ACTAS y CONSOLIDADOS dentro del MISMO Excel usando una PLANTILLA modelo.
# - No reemplaza tu RESULTADOS/RESUMEN: los agrega al archivo final.
# - Respeta el layout de tu plantilla (solo llena desde la fila de datos).
#
# Cambios solicitados:
# ✅ Columna EXAMEN: ahora se coloca el MES del examen (EJ: DICIEMBRE).
# ✅ Columna MODALIDAD: ahora se coloca "VIRTUAL".
# ✅ CONSOLIDADO: si la hoja tiene 3 bloques (A/B/C), llena cada bloque con su área.

from io import BytesIO
from datetime import datetime
import pandas as pd
import openpyxl


# -----------------------------
# Helpers de normalización
# -----------------------------
_MESES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}

def _mes_es(dt: datetime) -> str:
    return _MESES_ES.get(dt.month, "")

def _norm_dni(v) -> str:
    s = "" if pd.isna(v) else str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        return ""
    return digits.zfill(8)

def _area_letter(v) -> str:
    """
    Convierte valores como "A", "Área A – Ingenierías", "a" => "A".
    """
    s = "" if pd.isna(v) else str(v).strip().upper()
    if not s:
        return ""
    if s in ("A", "B", "C"):
        return s
    # si viene tipo "ÁREA A ..." tomamos la primera letra válida
    for ch in s:
        if ch in ("A", "B", "C"):
            return ch
    return ""

def _sede_key(s: str) -> str:
    s = (s or "").upper()
    if "CHINCHA" in s:
        return "CHINCHA"
    if "ICA" in s:
        return "ICA"
    return "CHINCHA"

def _clear_range(ws, r1: int, r2: int, c1: int = 1, c2: int = None):
    if c2 is None:
        c2 = ws.max_column
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).value = None

def _dump_df_values(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

def _find_rows_with_text(ws, text: str):
    text_u = text.upper()
    hits = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and text_u in v.upper():
                hits.add(r)
    return sorted(hits)

def _detect_block_area(ws, header_row: int, area_col: int = 8) -> str:
    """
    En tu plantilla de CONSOLIDADO, el bloque suele tener letras A/B/C en la columna 'AREA'
    en las primeras filas de datos. Detectamos eso para no depender del orden.
    """
    for r in range(header_row + 1, min(header_row + 15, ws.max_row + 1)):
        v = ws.cell(r, area_col).value
        letter = _area_letter(v)
        if letter in ("A", "B", "C"):
            return letter
    return ""


# -----------------------------
# Builder principal
# -----------------------------
def build_excel_final_con_actas(
    modelo_path: str,
    generated_excel_bytes: bytes,
    exam_date: datetime,
    exam_label: str = "EXAMEN ORDINARIO",  # (no se usa en columna EXAMEN; se mantiene por compatibilidad)
    output_add_resultados_resumen: bool = True,
    modalidad_value: str = "VIRTUAL",
) -> bytes:
    """
    Recibe el Excel ya generado por tu core (RESULTADOS + RESUMEN)
    y devuelve un único Excel final basado en la plantilla del modelo,
    rellenando:
      - Acta_Final_* (por sede)
      - Consolidado_* (por sede, y si tiene 3 bloques A/B/C los llena por área)
    y opcionalmente agregando RESULTADOS/RESUMEN al mismo archivo final.
    """

    # 1) leer Excel generado
    gen_xlsx = pd.ExcelFile(BytesIO(generated_excel_bytes))
    if "RESULTADOS" not in gen_xlsx.sheet_names or "RESUMEN" not in gen_xlsx.sheet_names:
        raise RuntimeError(f"El Excel generado no tiene RESULTADOS/RESUMEN. Hojas: {gen_xlsx.sheet_names}")

    df_res = pd.read_excel(gen_xlsx, sheet_name="RESULTADOS")
    df_sum = pd.read_excel(gen_xlsx, sheet_name="RESUMEN")

    # 2) validar columnas mínimas
    if "Numero de DNI" not in df_res.columns:
        raise RuntimeError("RESULTADOS no tiene la columna 'Numero de DNI'.")
    if "DNI" not in df_sum.columns:
        raise RuntimeError("RESUMEN no tiene la columna 'DNI'.")

    # 3) normalizar DNI para join
    df_res["_dni_norm"] = df_res["Numero de DNI"].apply(_norm_dni)
    df_sum["_dni_norm"] = df_sum["DNI"].apply(_norm_dni)

    # 4) armar df_small desde RESULTADOS
    cod_col = None
    for cand in ["Código de Matrícula", "Codigo de Matricula", "Código de Estudiante", "Codigo de Estudiante"]:
        if cand in df_res.columns:
            cod_col = cand
            break

    df_small = df_res[["_dni_norm"]].copy()
    df_small["APELLIDOS"] = df_res["Apellido(s)"] if "Apellido(s)" in df_res.columns else ""
    df_small["NOMBRES"] = df_res["Nombre"] if "Nombre" in df_res.columns else ""
    df_small["CORREO"] = df_res["Dirección de correo"] if "Dirección de correo" in df_res.columns else ""
    df_small["CODIGO"] = df_res[cod_col] if cod_col else ""

    base = df_sum.merge(df_small, on="_dni_norm", how="left")

    # columnas claves
    if "Sede o Filial" not in base.columns:
        raise RuntimeError("RESUMEN no tiene la columna 'Sede o Filial'.")
    if "Programa Académico" not in base.columns:
        raise RuntimeError("RESUMEN no tiene la columna 'Programa Académico'.")

    # normalizar sede y área
    base["SEDE_KEY"] = base["Sede o Filial"].astype(str).apply(_sede_key)
    if "Área" in base.columns:
        base["_AREA_LETTER"] = base["Área"].apply(_area_letter)
    else:
        base["_AREA_LETTER"] = ""

    # 5) abrir plantilla
    wb = openpyxl.load_workbook(modelo_path)

    # 6) mapeo de hojas: tolerante (por si cambias nombres)
    #    Prioridad: si existen nombres exactos, usa esos. Si no, busca por contiene "acta" / "consolid"
    def _find_sheet_by_contains(needle: str):
        needle = needle.lower()
        for s in wb.sheetnames:
            if needle in s.lower():
                return s
        return None

    # Actas por sede (si existen)
    acta_ch = "Acta_Final_Chincha" if "Acta_Final_Chincha" in wb.sheetnames else _find_sheet_by_contains("acta")
    cons_ch = "Consolidado_Chincha" if "Consolidado_Chincha" in wb.sheetnames else _find_sheet_by_contains("consol")

    # Si tu modelo tiene ICA también, se llenan; si no, no se rompe.
    acta_ica = "Acta_Final_Ica" if "Acta_Final_Ica" in wb.sheetnames else None
    cons_ica = "Consolidado_Ica" if "Consolidado_Ica" in wb.sheetnames else None

    if not acta_ch:
        raise RuntimeError(f"No pude detectar una hoja plantilla 'acta' dentro del modelo. Hojas: {wb.sheetnames}")
    if not cons_ch:
        raise RuntimeError(f"No pude detectar una hoja plantilla 'consolidado' dentro del modelo. Hojas: {wb.sheetnames}")

    # -----------------------------
    # Llenado "fila por fila"
    # -----------------------------
    # Layout por posición (1-based):
    #  1 N°, 2 APELLIDOS, 3 NOMBRES, 4 DNI, 5 CODIGO, 6 TELEFONO, 7 CORREO,
    #  8 AREA, 9 CARRERA, 10 SEDE DE ESTUDIO, 11 ASISTENCIA,
    #  12 COM, 13 %COM, 14 HAB, 15 %HAB, 16 MAT, 17 %MAT, 18 CTA, 19 %CTA,
    #  20 TOTAL, 21 %TOTAL, 22 CONDICIÓN, 23 MODALIDAD, 24 FECHA, 25 EXAMEN,
    #  26 MODALIDAD DE INGRESO (si existe)
    def _fill_rows(ws, df: pd.DataFrame, start_row: int, end_row: int):
        # limpiar data (mantener cabecera)
        _clear_range(ws, start_row, end_row, c1=1, c2=ws.max_column)

        # orden sugerido
        if "Programa Académico" in df.columns and "DNI" in df.columns:
            df = df.sort_values(["Programa Académico", "DNI"], kind="mergesort")

        mes = _mes_es(exam_date)

        for i in range(len(df)):
            rr = df.iloc[i]
            row_idx = start_row + i
            if row_idx > end_row:
                break  # no desborda el bloque

            ws.cell(row_idx, 1).value = i + 1
            ws.cell(row_idx, 2).value = rr.get("APELLIDOS", "")
            ws.cell(row_idx, 3).value = rr.get("NOMBRES", "")
            ws.cell(row_idx, 4).value = rr.get("DNI", "")
            ws.cell(row_idx, 5).value = rr.get("CODIGO", "") or ""
            ws.cell(row_idx, 6).value = ""  # TELEFONO
            ws.cell(row_idx, 7).value = rr.get("CORREO", "") or ""
            ws.cell(row_idx, 8).value = rr.get("Área", "") if "Área" in rr.index else ""
            ws.cell(row_idx, 9).value = rr.get("Programa Académico", "") or ""
            ws.cell(row_idx, 10).value = rr.get("Sede o Filial", "") or ""
            ws.cell(row_idx, 11).value = rr.get("Asistencia", "") if "Asistencia" in rr.index else ""

            ws.cell(row_idx, 12).value = rr.get("COMUNICACIÓN", "") if "COMUNICACIÓN" in rr.index else ""
            ws.cell(row_idx, 13).value = rr.get("% (COM)", "") if "% (COM)" in rr.index else ""
            ws.cell(row_idx, 14).value = rr.get("HABILIDADES COMUNICATIVAS", "") if "HABILIDADES COMUNICATIVAS" in rr.index else ""
            ws.cell(row_idx, 15).value = rr.get("% (HAB)", "") if "% (HAB)" in rr.index else ""
            ws.cell(row_idx, 16).value = rr.get("MATEMÁTICA", "") if "MATEMÁTICA" in rr.index else ""
            ws.cell(row_idx, 17).value = rr.get("% (MAT)", "") if "% (MAT)" in rr.index else ""
            ws.cell(row_idx, 18).value = rr.get("CTA/CCSS", "") if "CTA/CCSS" in rr.index else ""
            ws.cell(row_idx, 19).value = rr.get("% (CTA/CCSS)", "") if "% (CTA/CCSS)" in rr.index else ""
            ws.cell(row_idx, 20).value = rr.get("TOTAL", "") if "TOTAL" in rr.index else ""
            ws.cell(row_idx, 21).value = rr.get("%_TOTAL", "") if "%_TOTAL" in rr.index else ""
            ws.cell(row_idx, 22).value = rr.get("CONDICIÓN", "") if "CONDICIÓN" in rr.index else ""

            ws.cell(row_idx, 23).value = modalidad_value  # ✅ VIRTUAL
            ws.cell(row_idx, 24).value = exam_date.date()
            ws.cell(row_idx, 25).value = mes  # ✅ MES (DICIEMBRE, etc.)

            if ws.max_column >= 26:
                ws.cell(row_idx, 26).value = ""  # modalidad ingreso

    # -----------------------------
    # 7) ACTA: llenar por sede (una sola tabla desde fila 2 hasta fin)
    # -----------------------------
    def fill_acta(sheet_name: str, sede_key: str):
        ws = wb[sheet_name]
        df_sede = base[base["SEDE_KEY"] == sede_key].copy()
        start = 2
        end = ws.max_row
        _fill_rows(ws, df_sede, start_row=start, end_row=end)

    fill_acta(acta_ch, "CHINCHA")
    if acta_ica and acta_ica in wb.sheetnames:
        fill_acta(acta_ica, "ICA")

    # -----------------------------
    # 8) CONSOLIDADO: por sede. Si tiene 3 bloques, llenar cada bloque por AREA.
    # -----------------------------
    def fill_consolidado(sheet_name: str, sede_key: str):
        ws = wb[sheet_name]
        df_sede = base[base["SEDE_KEY"] == sede_key].copy()

        header_rows = _find_rows_with_text(ws, "APELLIDOS")
        if not header_rows:
            # fallback: si no detecta bloques, lo trata como tabla simple desde fila 2
            _fill_rows(ws, df_sede, start_row=2, end_row=ws.max_row)
            return

        # para cada bloque detectado
        for idx, hr in enumerate(header_rows):
            start_row = hr + 1

            # end_row: hasta antes del siguiente header o hasta max_row
            end_row = (header_rows[idx + 1] - 1) if (idx + 1 < len(header_rows)) else ws.max_row

            # detectar área del bloque
            area_blk = _detect_block_area(ws, hr, area_col=8)  # col 8 = AREA
            if not area_blk:
                # fallback por orden si no lo detecta
                area_blk = ["A", "B", "C"][idx] if idx < 3 else ""

            df_area = df_sede.copy()
            if area_blk in ("A", "B", "C"):
                df_area = df_area[df_area["_AREA_LETTER"] == area_blk].copy()

            _fill_rows(ws, df_area, start_row=start_row, end_row=end_row)

    fill_consolidado(cons_ch, "CHINCHA")
    if cons_ica and cons_ica in wb.sheetnames:
        fill_consolidado(cons_ica, "ICA")

    # 9) (Opcional) agregar RESULTADOS y RESUMEN
    if output_add_resultados_resumen:
        if "RESULTADOS" in wb.sheetnames:
            del wb["RESULTADOS"]
        if "RESUMEN" in wb.sheetnames:
            del wb["RESUMEN"]

        ws_r = wb.create_sheet("RESULTADOS", 0)
        ws_s = wb.create_sheet("RESUMEN", 1)

        _dump_df_values(ws_r, df_res.drop(columns=["_dni_norm"], errors="ignore"))
        _dump_df_values(ws_s, df_sum.drop(columns=["_dni_norm"], errors="ignore"))

    # 10) guardar
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
